"""
Excel log hovorů + zápis výsledku zpět do SQLite databáze obcí.
Excel je pro lidské čtení / export, SQLite je zdroj pravdy pro stav kampaně.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

import database
from config import settings

HEADERS = [
    "Obec", "Kraj", "Telefon", "Email", "Starosta/ka (zjištěno)",
    "Datum hovoru", "Čas hovoru", "Délka (min)", "Výsledek",
    "Termín schůzky", "Místo schůzky", "Zpětné volání", "Poznámka", "Přepis hovoru",
]

COL_WIDTHS = [20, 16, 18, 28, 22, 14, 10, 11, 22, 18, 22, 14, 35, 60]

OUTCOME_COLORS = {
    "meeting_agreed":     "0a2e0a",
    "not_interested":     "2e0a0a",
    "callback_requested": "2a200a",
    "send_email":         "0a152e",
    "voicemail":          "1a0a2e",
    "no_answer":          "1a1a1a",
    "wrong_person":       "201a0a",
    "ongoing":            "141414",
}

OUTCOME_LABELS = {
    "meeting_agreed":     "✅ Schůzka domluvena",
    "not_interested":     "❌ Nezájem",
    "callback_requested": "📞 Zpětné volání",
    "send_email":         "📧 Zaslat email",
    "voicemail":          "📬 Vzkaz/záznamník",
    "no_answer":          "🔇 Nedostupný",
    "wrong_person":       "🔁 Špatná osoba",
    "ongoing":            "⏳ Probíhá",
}


def _workbook() -> openpyxl.Workbook:
    path = Path(settings.EXCEL_LOG_FILE)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return openpyxl.load_workbook(path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hovory"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    hfill = PatternFill(start_color="0f1114", end_color="0f1114", fill_type="solid")
    hfont = Font(bold=True, color="d4a017", size=11)

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)
    return wb


def log_call(
    municipality_row: dict,
    outcome: str,
    call_duration_sec: int = 0,
    mayor_name: str = None,
    meeting_date: str = None,
    meeting_time: str = None,
    meeting_place: str = None,
    callback_date: str = None,
    notes: str = "",
    transcript: str = "",
) -> None:
    """Zapíše hovor do Excelu A zaktualizuje stav v SQLite databázi."""
    wb = _workbook()
    ws = wb.active
    now = datetime.now()

    meeting_str = f"{meeting_date} {meeting_time or ''}".strip() if meeting_date else ""
    place_str = meeting_place or (f"Obecní úřad {municipality_row.get('name', '')}" if meeting_date else "")

    row = [
        municipality_row.get("name", ""),
        municipality_row.get("kraj", ""),
        municipality_row.get("phone", ""),
        municipality_row.get("email", ""),
        mayor_name or "",
        now.strftime("%d.%m.%Y"),
        now.strftime("%H:%M"),
        round(call_duration_sec / 60, 1),
        OUTCOME_LABELS.get(outcome, outcome),
        meeting_str,
        place_str,
        callback_date or "",
        notes,
        transcript,
    ]

    color = OUTCOME_COLORS.get(outcome, "141414")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font_color = "22c55e" if outcome == "meeting_agreed" else "e8ecf4"

    idx = ws.max_row + 1
    for col, value in enumerate(row, 1):
        cell = ws.cell(row=idx, column=col, value=value)
        cell.fill = fill
        cell.font = Font(color=font_color, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=(col >= len(row) - 1))

    wb.save(settings.EXCEL_LOG_FILE)

    # Zápis stavu zpět do SQLite – zdroj pravdy pro "už voláno"
    database.update_call_result(
        municipality_id=municipality_row["id"],
        outcome=outcome,
        notes=notes,
        callback_date=callback_date,
        mayor_name=mayor_name,
    )


def get_log_as_rows() -> list[dict]:
    path = Path(settings.EXCEL_LOG_FILE)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append(dict(zip(HEADERS, row)))
    return rows
